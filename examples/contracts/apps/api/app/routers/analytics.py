"""Расширенная аналитика по договорам: срезы по продуктам/странам/менеджерам,
время согласования, возраст pending. admin/director/lawyer — все; manager — свои.

Эпик 6 MVP: + конверсия воронки (`/funnel/{pipeline_id}`) + forecast выручки
(`/forecast?pipeline_id=...`).

DEALS 2.0 Ф3: отчёты по сделкам
- /analytics/deals/awaiting-payment — сделки в подстатусе await_payment
- /analytics/deals/paid           — сделки в подстатусе paid
- /analytics/deals/hot            — сделки в этапе hot (прогноз дожатия)
"""
from __future__ import annotations

from collections import Counter
from datetime import UTC, date, datetime, timedelta
from decimal import Decimal
from io import BytesIO
from typing import Annotated

from fastapi import APIRouter, Depends, HTTPException, Query, Response
from sqlalchemy import Date, DateTime, and_, false, func, or_, true
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select

from app.db import get_session
from app.deps import CurrentUser, DirectorOrAdmin
from app.models import (
    Activity,
    Approval,
    ApprovalDecision,
    ClientSubscription,
    Company,
    Contract,
    ContractPayment,
    ContractStatus,
    Counterparty,
    Deal,
    Pipeline,
    PipelineStage,
    SubscriptionStageHistory,
    User,
)
from app.services.analytics import (
    build_kpi_trends_from_aggregates,
    build_xlsx,
    compute_forecast_revenue,
    compute_funnel_metrics,
    weekly_buckets_from_counts,
)
from app.services.contract_status import group_codes_in_order as contract_status_group_codes
from app.services.contract_status import primary_group as contract_primary_group
from app.services.analytics_deals import (
    TASK_LIKE_ACTIVITY_KINDS,
    aggregate_awaiting_payment,
    compute_hot_forecast,
    compute_paid_summary,
)
from app.services.cohort_analytics import (
    build_cohort_data,
    build_cohort_xlsx_rows,
    compute_monthly_churn_rate,
    compute_projected_ltv,
    compute_cohort_matrix,
)

_XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

router = APIRouter(prefix="/analytics", tags=["analytics"])

# Wave 2a: «pending» = договор всё ещё ждёт решения по согласованию.
# Включаем needs_rework (вернули на доработку — цикл ещё активен, автор работает).
# rejected НЕ включаем: отклонён = разрешённое (terminal) состояние попытки, не
# «ждёт». Используется в pending_count / pending_avg_age_days.
_PENDING = (
    ContractStatus.submitted,
    ContractStatus.in_review,
    ContractStatus.needs_rework,
)


def _avg_epoch_days(start_col, end_col):
    """SQL-выражение: средняя разница (end - start) в днях.

    Эквивалент services.analytics.avg_days: per-row (end-start) переводится в
    дни (доля суток), затем avg по строкам. Округление до 1 знака делаем в
    Python после fetch (чтобы совпадало с round(..., 1)). NULL-пары PG отбросит
    из avg автоматически (как и avg_days пропускает None).
    """
    return func.avg(
        func.extract("epoch", end_col - start_col) / 86400.0
    )


def _round1(v) -> float:
    """Округлить SQL-avg (Decimal|float|None) до 1 знака как avg_days."""
    return round(float(v), 1) if v is not None else 0.0


async def _contracts_avg_time_to_approve(
    session: AsyncSession,
    base_where,
    start_window: datetime | None = None,
    end_window: datetime | None = None,
) -> float:
    """avg_time_to_approve_days через сфокусированный подзапрос (без full-table).

    Логика идентична старой in-memory:
      per-contract start = min(approval.created_at) по ВСЕМ попыткам;
            end   = max(approval.decided_at) по approved-попыткам;
      учитываем только договоры, у которых есть хотя бы один approved-голос;
      затем avg(end - start) в днях, round(.,1).

    Реализация: подзапрос группирует approvals по contract_id и считает
    min(created_at) и max(decided_at) FILTER (WHERE approved). HAVING отсекает
    контракты без approved. base_where ограничивает множество договоров
    (scope + since_days), start/end_window — опциональное окно по created_at
    договора (для trend-периодов).
    """
    cstart = func.min(Approval.created_at)
    cend = func.max(Approval.decided_at).filter(
        Approval.decision == ApprovalDecision.approved,
        Approval.decided_at.isnot(None),
    )
    has_approved = func.count(Approval.id).filter(
        Approval.decision == ApprovalDecision.approved,
        Approval.decided_at.isnot(None),
    )
    per_contract = (
        select(
            Approval.contract_id.label("cid"),
            cstart.label("start_at"),
            cend.label("end_at"),
        )
        .join(Contract, Contract.id == Approval.contract_id)
        .where(base_where)
    )
    if start_window is not None:
        per_contract = per_contract.where(Contract.created_at >= start_window)
    if end_window is not None:
        per_contract = per_contract.where(Contract.created_at < end_window)
    per_contract = (
        per_contract.group_by(Approval.contract_id)
        .having(has_approved > 0)
        .subquery()
    )
    avg_stmt = select(
        _avg_epoch_days(per_contract.c.start_at, per_contract.c.end_at)
    )
    return _round1((await session.execute(avg_stmt)).scalar_one())


@router.get("/contracts")
async def contracts_analytics(
    current_user: CurrentUser,
    session: Annotated[AsyncSession, Depends(get_session)],
    since_days: int | None = None,
):
    """Срезы по договорам через SQL GROUP BY / aggregate (без загрузки таблицы).

    P2 (audit B4): раньше эндпоинт грузил ВСЮ таблицу contracts + все approvals
    в Python и агрегировал Counter/defaultdict-циклами — бомба latency/memory на
    10k+ строк. Теперь все срезы считаются в SQL; форма ответа не изменилась,
    role-scope (manager → свои) применён в WHERE.
    """
    from app.deps import _SCOPE_ELEVATED_ROLES

    now = datetime.now(UTC)

    # Базовый WHERE: scope (fail-CLOSED whitelist) + since_days.
    # Используем _SCOPE_ELEVATED_ROLES (admin/director/lawyer) — видят всё.
    # Любая другая роль (manager, accountant, cfo, ...) — только свои договоры.
    # Применяется во ВСЕХ агрегатах (в SQL), не после fetch.
    base_conds = []
    if since_days:
        base_conds.append(Contract.created_at >= now - timedelta(days=since_days))
    if current_user.role not in _SCOPE_ELEVATED_ROLES:
        base_conds.append(Contract.author_user_id == current_user.id)
    base_where = and_(true(), *base_conds)

    # total
    total = (await session.execute(
        select(func.count(Contract.id)).where(base_where)
    )).scalar_one()

    # by_product — GROUP BY product_code
    by_product = {
        row[0]: row[1]
        for row in (await session.execute(
            select(Contract.product_code, func.count(Contract.id))
            .where(base_where).group_by(Contract.product_code)
        )).all()
    }

    # by_country — GROUP BY upper(coalesce(country, '—'))
    country_expr = func.upper(func.coalesce(func.nullif(Contract.country_code, ""), "—"))
    by_country = {
        (row[0] or "—"): row[1]
        for row in (await session.execute(
            select(country_expr, func.count(Contract.id))
            .where(base_where).group_by(country_expr)
        )).all()
    }

    # by_status — GROUP BY status. Возвращаем .value enum'а как ключ.
    status_rows = (await session.execute(
        select(Contract.status, func.count(Contract.id))
        .where(base_where).group_by(Contract.status)
    )).all()
    by_status = {st.value: cnt for st, cnt in status_rows}

    # by_status_group — те же per-status счётчики, свёрнутые в 4 первичные группы.
    # Post-step фиксированного размера (<= число статусов), не per-row — это OK.
    by_status_group: Counter = Counter()
    for code in contract_status_group_codes():
        by_status_group[code] = 0
    for st, cnt in status_rows:
        by_status_group[contract_primary_group(st)["code"]] += cnt

    # by_manager — GROUP BY author_user_id + join User.full_name, order count desc
    # (зеркалит старый Counter.most_common()). Имя отсутствующего юзера → «—».
    manager_rows = (await session.execute(
        select(
            func.coalesce(User.full_name, "—").label("name"),
            func.count(Contract.id).label("cnt"),
        )
        .select_from(Contract)
        .outerjoin(User, User.id == Contract.author_user_id)
        .where(base_where)
        .group_by(User.full_name)
        .order_by(func.count(Contract.id).desc())
    )).all()
    by_manager = {name: cnt for name, cnt in manager_rows}

    # pending_count + pending_avg_age_days (created_at → now) — SQL.
    pending_where = and_(base_where, Contract.status.in_(_PENDING))
    pending_count, pending_avg_age = (await session.execute(
        select(
            func.count(Contract.id),
            _avg_epoch_days(Contract.created_at, func.cast(now, DateTime(timezone=True))),
        ).where(pending_where)
    )).one()
    pending_avg_age = _round1(pending_avg_age)

    # avg_cycle_days (created_at → signed_at) — SQL avg over signed.
    avg_cycle = _round1((await session.execute(
        select(_avg_epoch_days(Contract.created_at, Contract.signed_at))
        .where(and_(base_where, Contract.signed_at.isnot(None)))
    )).scalar_one())

    # avg_time_to_approve_days — сфокусированный подзапрос (см. helper).
    avg_time_to_approve = await _contracts_avg_time_to_approve(session, base_where)
    approved_sample = (await session.execute(
        select(func.count())
        .select_from(
            select(Approval.contract_id)
            .join(Contract, Contract.id == Approval.contract_id)
            .where(base_where)
            .group_by(Approval.contract_id)
            .having(
                func.count(Approval.id).filter(
                    Approval.decision == ApprovalDecision.approved,
                    Approval.decided_at.isnot(None),
                ) > 0
            )
            .subquery()
        )
    )).scalar_one()

    # ---- Design v2: sparkline (8 недель) + trend_pct — всё из SQL-агрегатов ----
    sparkline_weeks = 8
    window_start = (now - timedelta(weeks=sparkline_weeks)).date()
    # Дневные счётчики новых договоров за окно sparkline (GROUP BY дата).
    day_col = func.cast(Contract.created_at, Date)
    daily_rows = (await session.execute(
        select(day_col, func.count(Contract.id))
        .where(and_(base_where, day_col >= window_start))
        .group_by(day_col)
    )).all()
    weekly_counts = weekly_buckets_from_counts(
        [(d, c) for d, c in daily_rows], weeks=sparkline_weeks, ref_date=now.date()
    )

    # avg_cycle / avg_time_to_approve по 30-дневным окнам для trend_pct.
    cutoff_30 = now - timedelta(days=30)
    cutoff_60 = now - timedelta(days=60)

    async def _avg_cycle_window(start_w, end_w) -> float:
        conds = [base_where, Contract.signed_at.isnot(None), Contract.created_at >= start_w]
        if end_w is not None:
            conds.append(Contract.created_at < end_w)
        return _round1((await session.execute(
            select(_avg_epoch_days(Contract.created_at, Contract.signed_at))
            .where(and_(*conds))
        )).scalar_one())

    cycle_current = await _avg_cycle_window(cutoff_30, now)
    cycle_prev = await _avg_cycle_window(cutoff_60, cutoff_30)
    tta_current = await _contracts_avg_time_to_approve(session, base_where, cutoff_30, now)
    tta_prev = await _contracts_avg_time_to_approve(session, base_where, cutoff_60, cutoff_30)

    kpi_trends = build_kpi_trends_from_aggregates(
        weekly_counts=weekly_counts,
        cycle_current=cycle_current,
        cycle_prev=cycle_prev,
        tta_current=tta_current,
        tta_prev=tta_prev,
    )

    return {
        "total": total,
        "by_product": by_product,
        "by_country": by_country,
        "by_manager": by_manager,
        "by_status": by_status,
        # Wave 2a: те же договоры, сгруппированные по 4 первичным статус-группам.
        "by_status_group": dict(by_status_group),
        "pending_count": pending_count,
        "pending_avg_age_days": pending_avg_age,
        "avg_cycle_days": avg_cycle,           # создан → сделка проведена
        "avg_time_to_approve_days": avg_time_to_approve,  # отправлен → согласован
        "approved_sample": approved_sample,
        # Design v2: KPI sparkline + trend (Optional — фронт не падает если None).
        # total_sparkline: list[int] — 8 недель новых договоров (старые→новые).
        # total_trend_pct: float|null — % изменения count за 4w vs предыдущие 4w.
        # avg_cycle_trend_pct: float|null — % изм. avg_cycle_days (30d vs 30d).
        # avg_time_to_approve_trend_pct: float|null — % изм. avg_time_to_approve.
        "total_sparkline": kpi_trends["total_sparkline"],
        "total_trend_pct": kpi_trends["total_trend_pct"],
        "avg_cycle_trend_pct": kpi_trends["avg_cycle_trend_pct"],
        "avg_time_to_approve_trend_pct": kpi_trends["avg_time_to_approve_trend_pct"],
    }


def _d(x) -> str:
    return x.date().isoformat() if x else ""


@router.get("/contracts.xlsx")
async def export_contracts(current_user: CurrentUser, session: Annotated[AsyncSession, Depends(get_session)]):
    """Выгрузка реестра договоров в Excel.

    Scope: fail-CLOSED whitelist (_SCOPE_ELEVATED_ROLES: admin/director/lawyer) — видят всё.
    Любая другая роль (manager, accountant, cfo, ...) — только свои договоры.
    """
    from app.deps import scope_to_user
    stmt = select(Contract).order_by(Contract.created_at.desc())
    stmt = scope_to_user(stmt, Contract, current_user, "author_user_id")
    contracts = (await session.execute(stmt)).scalars().all()

    # CONTACTS 2.0 Ф4: имя клиента берём из Company (источник истины).
    # Fallback: если company_id пуст — через counterparty_id → Counterparty.
    company_ids_xlsx = {c.company_id for c in contracts if c.company_id}
    cp_ids_fallback = {c.counterparty_id for c in contracts if not c.company_id and c.counterparty_id}
    author_ids = {c.author_user_id for c in contracts}

    company_names: dict[int, str] = {}
    if company_ids_xlsx:
        company_names = {
            co.id: (co.name or co.legal_name or "")
            for co in (await session.execute(
                select(Company).where(Company.id.in_(company_ids_xlsx))
            )).scalars().all()
        }
    cp_names_fallback: dict[int, str] = {}
    if cp_ids_fallback:
        cp_names_fallback = {
            cp.id: (cp.name or "")
            for cp in (await session.execute(
                select(Counterparty).where(Counterparty.id.in_(cp_ids_fallback))
            )).scalars().all()
        }
    names = {u.id: u.full_name for u in (await session.execute(select(User).where(User.id.in_(author_ids)))).scalars().all()} if author_ids else {}

    def _client_name_contract(c: Contract) -> str:
        if c.company_id:
            return company_names.get(c.company_id, "")
        return cp_names_fallback.get(c.counterparty_id, "") if c.counterparty_id else ""

    headers = ["Номер", "Название", "Продукт", "Страна", "Город", "Статус", "Контрагент", "Автор", "Создан", "Подписан", "Сумма", "Валюта"]
    rows = [[
        c.number, c.title, c.product_code, (c.country_code or "").upper(), c.city, c.status.value,
        _client_name_contract(c), names.get(c.author_user_id, ""), _d(c.created_at), _d(c.signed_at),
        float(c.total) if c.total is not None else "", c.currency,
    ] for c in contracts]
    data = build_xlsx("Договоры", headers, rows)
    return Response(content=data, media_type=_XLSX_MEDIA, headers={"Content-Disposition": 'attachment; filename="contracts.xlsx"'})


@router.get("/registry.xlsx")
async def export_registry(current_user: CurrentUser, session: Annotated[AsyncSession, Depends(get_session)]):
    """Выгрузка реестра клиентов (CS) в Excel.

    Scope: зеркалит registry.py — scope_query(session, stmt, ClientSubscription, "subscription", user).
    Fail-CLOSED: elevated роли (admin/director/lawyer) видят весь реестр,
    остальные (manager, accountant, cfo, ...) — только свои подписки.
    """
    from app.routers.registry import _registry_rows
    from app.services.access_control import scope_query

    stmt = select(ClientSubscription).where(ClientSubscription.is_active.is_(True)).order_by(ClientSubscription.id)
    stmt = await scope_query(session, stmt, ClientSubscription, "subscription", current_user)
    subs = (await session.execute(stmt)).scalars().all()
    rows_data = await _registry_rows(session, list(subs))

    # CONTACTS 2.0 Ф3-B: город берём из Company (источник истины), fallback на Counterparty.
    company_ids_excel = {r.company_id for r in rows_data if r.company_id}
    cp_ids_excel = {r.counterparty_id for r in rows_data if not r.company_id}

    company_city: dict[int, str] = {}
    if company_ids_excel:
        company_city = {
            c.id: (c.city or "")
            for c in (await session.execute(
                select(Company).where(Company.id.in_(company_ids_excel))
            )).scalars().all()
        }
    cp_city: dict[int, str] = {}
    if cp_ids_excel:
        cp_city = {
            c.id: (c.city or "")
            for c in (await session.execute(
                select(Counterparty).where(Counterparty.id.in_(cp_ids_excel))
            )).scalars().all()
        }

    headers = ["Клиент", "Страна", "Город", "Платформа", "Регион", "Категория", "Статус", "Внедрение %",
               "Здоровье", "Тренд %", "Абонентка", "Валюта", "Тариф", "Ответственный", "Требует внимания"]
    rows = [[
        r.counterparty_name,
        (r.country_code or "").upper(),
        company_city.get(r.company_id, "") if r.company_id else cp_city.get(r.counterparty_id, ""),
        r.platform_name, r.region_name or "", r.category_code or "", r.status_code or "", r.impl_pct,
        r.health_tier or "", r.activity_trend_pct, r.fee_actual, r.fee_currency or "", r.tariff or "",
        r.sup_pm_name or r.am_name or "", ", ".join(r.attention),
    ] for r in rows_data]
    data = build_xlsx("Реестр клиентов", headers, rows)
    return Response(content=data, media_type=_XLSX_MEDIA, headers={"Content-Disposition": 'attachment; filename="registry.xlsx"'})


# ============ Конверсия воронки (Эпик 6 MVP) ============


@router.get("/funnel/{pipeline_id}")
async def funnel_conversion(
    pipeline_id: int,
    current_user: CurrentUser,
    session: Annotated[AsyncSession, Depends(get_session)],
):
    """Per-stage метрики для воронки (любой kind: sales/lifecycle/lead/renewal).

    Возвращает:
    {
        pipeline: {id, name, kind},
        stages: [{
            stage_id, stage_name, stage_code, sort_order, count,
            avg_days_in_stage, transition_to_next_pct,
            is_won, is_lost, probability
        }],
        total_active, total_won, total_lost
    }

    ВАЖНО: без DealStageHistory time-in-stage приближённый (используем
    Deal.updated_at — может перетереться при любой правке сделки, не только
    смене этапа). Точная метрика — в эпике 12 после добавления полноценного
    лога переходов (есть модель DealStageHistory, но не пишется автоматом).
    """
    pipe = (await session.execute(
        select(Pipeline).where(Pipeline.id == pipeline_id)
    )).scalar_one_or_none()
    if not pipe:
        raise HTTPException(404, "Воронка не найдена")

    stages = (await session.execute(
        select(PipelineStage)
        .where(PipelineStage.pipeline_id == pipeline_id)
        .order_by(PipelineStage.sort_order)
    )).scalars().all()
    # Scope: fail-CLOSED через scope_query (как в /deals board).
    # admin/director/lawyer видят все сделки; остальные (manager, accountant, cfo) — свои.
    deals_stmt = select(Deal).where(Deal.pipeline_id == pipeline_id)
    deals_stmt = await _apply_deal_scope(session, deals_stmt, current_user)
    deals = (await session.execute(deals_stmt)).scalars().all()

    stages_data = [
        {
            "id": s.id,
            "name": s.name,
            "code": s.code,
            "sort_order": s.sort_order,
            "is_won": s.is_won,
            "is_lost": s.is_lost,
        }
        for s in stages
    ]
    deals_data = [
        {
            "stage_id": d.stage_id,
            "updated_at": d.updated_at,
            "amount": d.amount,
        }
        for d in deals
    ]
    metrics = compute_funnel_metrics(stages_data, deals_data)

    total_active = sum(m["count"] for m in metrics if not m["is_won"] and not m["is_lost"])
    total_won = sum(m["count"] for m in metrics if m["is_won"])
    total_lost = sum(m["count"] for m in metrics if m["is_lost"])

    return {
        "pipeline": {"id": pipe.id, "name": pipe.name, "kind": pipe.kind},
        "stages": metrics,
        "total_active": total_active,
        "total_won": total_won,
        "total_lost": total_lost,
    }


@router.get("/forecast")
async def forecast_revenue(
    current_user: CurrentUser,
    session: Annotated[AsyncSession, Depends(get_session)],
    pipeline_id: int | None = Query(
        None,
        description=(
            "если не задан — берётся первая активная Pipeline.kind='sales'"
        ),
    ),
):
    """Простой forecast выручки по воронке (sales по умолчанию).

    Алгоритм: sum(active_count × avg_won_value × probability_by_stage).
    Probability — keyword-based heuristic на названии этапа (HOT=0.7, warm=0.4,
    Trial=0.5, ...). Полноценная конфигурация probability_by_stage — в
    Pipeline.config либо отдельная таблица SalesPlan (эпик 10).
    """
    if pipeline_id is None:
        # Дефолт: первая активная sales-воронка
        pipe = (await session.execute(
            select(Pipeline)
            .where(Pipeline.kind == "sales", Pipeline.is_active.is_(True))
            .order_by(Pipeline.id)
            .limit(1)
        )).scalar_one_or_none()
        if not pipe:
            raise HTTPException(404, "sales-воронка не найдена")
    else:
        pipe = (await session.execute(
            select(Pipeline).where(Pipeline.id == pipeline_id)
        )).scalar_one_or_none()
        if not pipe:
            raise HTTPException(404, "Воронка не найдена")

    stages = (await session.execute(
        select(PipelineStage)
        .where(PipelineStage.pipeline_id == pipe.id)
        .order_by(PipelineStage.sort_order)
    )).scalars().all()
    # Scope: fail-CLOSED через scope_query (как в /deals board).
    # admin/director/lawyer видят все сделки; остальные (manager, accountant, cfo) — свои.
    deals_stmt = select(Deal).where(Deal.pipeline_id == pipe.id)
    deals_stmt = await _apply_deal_scope(session, deals_stmt, current_user)
    deals = (await session.execute(deals_stmt)).scalars().all()

    stages_data = [
        {
            "id": s.id,
            "name": s.name,
            "is_won": s.is_won,
            "is_lost": s.is_lost,
            "sort_order": s.sort_order,
        }
        for s in stages
    ]
    deals_data = [
        {
            "stage_id": d.stage_id,
            "amount": (
                float(d.amount) if isinstance(d.amount, (int, float, Decimal))
                else None
            ),
            "currency": d.currency,
        }
        for d in deals
    ]
    result = compute_forecast_revenue(stages_data, deals_data)
    return {
        "pipeline": {"id": pipe.id, "name": pipe.name, "kind": pipe.kind},
        **result,
    }


# ============ Cohort Analytics (Эпик 22) ============


def _build_sub_dicts(
    subs: list[ClientSubscription],
    churn_dates: dict[int, datetime],
    stage_codes: dict[int, str],
) -> list[dict]:
    """Собрать список dict для cohort service functions.

    cohort_start_date = impl_start_date если есть, иначе created_at.
    churn_date = дата первого перехода в C0 (из subscription_stage_history),
    или None если подписка ещё активна.
    fee_actual берётся напрямую из ClientSubscription.
    """
    result = []
    for sub in subs:
        # Дата начала когорты: предпочитаем impl_start_date, fallback — created_at
        cohort_start = (
            sub.impl_start_date
            if sub.impl_start_date is not None
            else (sub.created_at.date() if sub.created_at else None)
        )
        if cohort_start is None:
            continue

        churn_dt = churn_dates.get(sub.id)
        churn_date = churn_dt.date() if churn_dt else None

        # Получаем code текущего stage для фильтра
        current_stage_code = stage_codes.get(sub.lifecycle_stage_id) if sub.lifecycle_stage_id else None

        result.append({
            "subscription_id": sub.id,
            "cohort_start_date": cohort_start,
            "churn_date": churn_date,
            "fee_actual": sub.fee_actual,
            "current_stage_code": current_stage_code,
            "is_active": sub.is_active,
        })
    return result


@router.get("/cohorts")
async def cohorts_analytics(
    current_user: CurrentUser,
    session: Annotated[AsyncSession, Depends(get_session)],
    cohort_type: str = Query("monthly", description="monthly или quarterly (MVP: только monthly)"),
    periods: int = Query(12, ge=1, le=36, description="Глубина анализа в месяцах"),
    product_code: str | None = Query(None, description="Фильтр по коду платформы (platform.code)"),
    country_code: str | None = Query(None, description="Фильтр по коду страны контрагента"),
):
    """Когортная матрица retention + LTV для подписок CS-реестра.

    Cohort = группа подписок, активированных в одном месяце (cohort_start_date).
    Retention = % подписок, не перешедших в C0 через N месяцев от starта.

    Источник «даты смерти»: subscription_stage_history (первый переход в C0).
    Если история пустая (до backfill) — используем текущий stage приближённо:
    stage_code='C0' → churn_date = subscription.updated_at.

    Response:
    {
        cohorts: [CohortData],          -- отсортировано по cohort_month
        matrix: {cohort_month: {offset: count}},
        retention_pct: {cohort_month: {offset: pct}},
        avg_ltv_per_cohort: {cohort_month: avg_ltv},
        total_avg_ltv: float,           -- среднее по всем когортам
        projected_ltv: float,           -- MRR / avg_churn_rate
        monthly_churn_rate: float,      -- средний месячный churn [0..1]
        current_mrr: float,             -- текущий MRR (сумма fee_actual активных)
    }
    """
    # Загружаем подписки через scope_query — fail-CLOSED whitelist.
    # admin/director/lawyer видят весь реестр; остальные (manager, accountant, cfo) — свои.
    # scope_query использует owner_user_id ClientSubscription через visibility_settings.
    from app.services.access_control import scope_query as _sq
    stmt = select(ClientSubscription)
    stmt = await _sq(session, stmt, ClientSubscription, "subscription", current_user)

    # CONTACTS 2.0 Ф4: фильтр по стране — сначала через Company (источник истины),
    # fallback на Counterparty для подписок без company_id.
    if country_code:
        cc_upper = country_code.upper()
        # Компании с нужной страной (Company.country || Company.country_code)
        co_ids_result = await session.execute(
            select(Company.id).where(
                func.upper(func.coalesce(Company.country, Company.country_code)) == cc_upper
            )
        )
        co_ids = [r[0] for r in co_ids_result]

        # Counterparty-fallback для подписок, у которых нет company_id
        cp_ids_result = await session.execute(
            select(Counterparty.id).where(
                func.upper(Counterparty.country_code) == cc_upper
            )
        )
        cp_ids = [r[0] for r in cp_ids_result]

        if not co_ids and not cp_ids:
            return _empty_cohort_response()

        # Подписка подходит, если её company_id в co_ids ИЛИ (нет company_id И counterparty_id в cp_ids)
        stmt = stmt.where(
            or_(
                ClientSubscription.company_id.in_(co_ids) if co_ids else false(),
                and_(
                    ClientSubscription.company_id.is_(None),
                    ClientSubscription.counterparty_id.in_(cp_ids) if cp_ids else false(),
                ),
            )
        )

    subs = (await session.execute(stmt)).scalars().all()
    if not subs:
        return _empty_cohort_response()

    # Загружаем stage codes для всех lifecycle_stage_id одним запросом
    stage_ids = {s.lifecycle_stage_id for s in subs if s.lifecycle_stage_id}
    stage_codes: dict[int, str] = {}
    if stage_ids:
        stages_rows = (await session.execute(
            select(PipelineStage).where(PipelineStage.id.in_(stage_ids))
        )).scalars().all()
        stage_codes = {s.id: s.code for s in stages_rows}

    # Фильтр по product_code (platform) — PipelineStage не хранит код платформы,
    # платформа хранится в ClientSubscription.platform_id. Но product_code в задаче
    # это platform.code — загружаем маппинг.
    if product_code:
        from app.models import Platform
        platform_result = await session.execute(
            select(Platform.id).where(Platform.code == product_code)
        )
        platform_ids = [r[0] for r in platform_result]
        if not platform_ids:
            return _empty_cohort_response()
        subs = [s for s in subs if s.platform_id in platform_ids]
        if not subs:
            return _empty_cohort_response()

    sub_ids = [s.id for s in subs]

    # Загружаем даты первого перехода в C0 из истории (один запрос, агрегация в Python)
    # Если history пуста — fallback на current stage
    churn_dates: dict[int, datetime] = {}
    if sub_ids:
        history_rows = (await session.execute(
            select(SubscriptionStageHistory).where(
                SubscriptionStageHistory.subscription_id.in_(sub_ids),
                SubscriptionStageHistory.to_stage_code == "C0",
            ).order_by(SubscriptionStageHistory.changed_at)
        )).scalars().all()

        # Берём самую раннюю дату перехода в C0 для каждой подписки
        for h in history_rows:
            if h.subscription_id not in churn_dates:
                churn_dates[h.subscription_id] = h.changed_at

    # Fallback: если нет history но stage_code == 'C0' — используем updated_at
    for sub in subs:
        if sub.id in churn_dates:
            continue
        stage_code = stage_codes.get(sub.lifecycle_stage_id) if sub.lifecycle_stage_id else None
        if stage_code == "C0" and sub.updated_at:
            churn_dates[sub.id] = sub.updated_at

    # Строим структуры для сервисных функций
    sub_dicts = _build_sub_dicts(subs, churn_dates, stage_codes)

    # Когортная матрица + retention
    cohorts_data = build_cohort_data(sub_dicts, periods)
    matrix = compute_cohort_matrix(sub_dicts, periods)
    from app.services.cohort_analytics import compute_retention_percent, compute_avg_ltv_per_cohort
    retention_pct = compute_retention_percent(matrix)
    avg_ltv_per_cohort = compute_avg_ltv_per_cohort(sub_dicts)

    # MRR: сумма fee_actual у активных подписок (stage != C0 и is_active=True)
    current_mrr = sum(
        float(s.fee_actual)
        for s in subs
        if s.is_active and s.fee_actual is not None
        and stage_codes.get(s.lifecycle_stage_id) != "C0"
    )

    monthly_churn_rate = compute_monthly_churn_rate(matrix)
    projected_ltv = compute_projected_ltv(current_mrr, monthly_churn_rate)

    total_avg_ltv = (
        round(sum(avg_ltv_per_cohort.values()) / len(avg_ltv_per_cohort), 2)
        if avg_ltv_per_cohort else 0.0
    )

    return {
        "cohorts": cohorts_data,
        "matrix": {k: dict(v) for k, v in matrix.items()},
        "retention_pct": {k: dict(v) for k, v in retention_pct.items()},
        "avg_ltv_per_cohort": avg_ltv_per_cohort,
        "total_avg_ltv": total_avg_ltv,
        "projected_ltv": projected_ltv,
        "monthly_churn_rate": monthly_churn_rate,
        "current_mrr": current_mrr,
    }


def _empty_cohort_response() -> dict:
    return {
        "cohorts": [],
        "matrix": {},
        "retention_pct": {},
        "avg_ltv_per_cohort": {},
        "total_avg_ltv": 0.0,
        "projected_ltv": 0.0,
        "monthly_churn_rate": 0.0,
        "current_mrr": 0.0,
    }


@router.get("/cohorts/{cohort_month}/members")
async def cohort_members(
    cohort_month: str,
    current_user: CurrentUser,
    session: Annotated[AsyncSession, Depends(get_session)],
):
    """Список подписок-участников конкретной когорты (cohort_month = 'YYYY-MM').

    Response: список объектов с базовыми данными подписки + статус retention.
    Используется в drill-down при клике на строку матрицы.
    """
    # Парсим cohort_month
    try:
        year, month = int(cohort_month[:4]), int(cohort_month[5:7])
    except (ValueError, IndexError):
        raise HTTPException(400, "cohort_month должен быть в формате YYYY-MM")

    from datetime import date
    cohort_start = date(year, month, 1)
    # Следующий месяц = конец диапазона для фильтра
    next_year = year + (month // 12)
    next_month = (month % 12) + 1
    cohort_end = date(next_year, next_month, 1)

    # Подписки с cohort_start в диапазоне [cohort_start, cohort_end)
    # impl_start_date предпочтительна, fallback на created_at
    # Scope: fail-CLOSED через scope_query (зеркал /registry).
    from app.services.access_control import scope_query as _sq
    stmt = select(ClientSubscription)
    stmt = await _sq(session, stmt, ClientSubscription, "subscription", current_user)
    all_subs = (await session.execute(stmt)).scalars().all()

    # Фильтрация по cohort_month в Python (impl_start_date || created_at)
    cohort_subs = []
    for sub in all_subs:
        csd = sub.impl_start_date or (sub.created_at.date() if sub.created_at else None)
        if csd and cohort_start <= csd < cohort_end:
            cohort_subs.append(sub)

    if not cohort_subs:
        return {"cohort_month": cohort_month, "members": [], "count": 0}

    sub_ids = [s.id for s in cohort_subs]

    # stage codes
    stage_ids = {s.lifecycle_stage_id for s in cohort_subs if s.lifecycle_stage_id}
    stage_codes: dict[int, str] = {}
    stage_names: dict[int, str] = {}
    if stage_ids:
        stages_rows = (await session.execute(
            select(PipelineStage).where(PipelineStage.id.in_(stage_ids))
        )).scalars().all()
        stage_codes = {s.id: s.code for s in stages_rows}
        stage_names = {s.id: s.name for s in stages_rows}

    # churn dates из history
    churn_dates: dict[int, datetime] = {}
    history_rows = (await session.execute(
        select(SubscriptionStageHistory).where(
            SubscriptionStageHistory.subscription_id.in_(sub_ids),
            SubscriptionStageHistory.to_stage_code == "C0",
        ).order_by(SubscriptionStageHistory.changed_at)
    )).scalars().all()
    for h in history_rows:
        if h.subscription_id not in churn_dates:
            churn_dates[h.subscription_id] = h.changed_at

    # CONTACTS 2.0 Ф4: имя клиента из Company, fallback на Counterparty.
    co_ids_members = {s.company_id for s in cohort_subs if s.company_id}
    cp_ids_fallback_members = {s.counterparty_id for s in cohort_subs if not s.company_id and s.counterparty_id}

    company_map: dict[int, str] = {}
    if co_ids_members:
        cos = (await session.execute(
            select(Company).where(Company.id.in_(co_ids_members))
        )).scalars().all()
        company_map = {c.id: (c.name or c.legal_name or "") for c in cos}

    cp_map: dict[int, str] = {}
    if cp_ids_fallback_members:
        cps_rows = (await session.execute(
            select(Counterparty).where(Counterparty.id.in_(cp_ids_fallback_members))
        )).scalars().all()
        cp_map = {c.id: (c.name or "") for c in cps_rows}

    def _member_client_name(sub: ClientSubscription) -> str:
        if sub.company_id:
            return company_map.get(sub.company_id, "")
        return cp_map.get(sub.counterparty_id, "") if sub.counterparty_id else ""

    members = []
    for sub in cohort_subs:
        stage_code = stage_codes.get(sub.lifecycle_stage_id) if sub.lifecycle_stage_id else None
        churn_dt = churn_dates.get(sub.id)
        # Fallback churn если нет history
        if churn_dt is None and stage_code == "C0" and sub.updated_at:
            churn_dt = sub.updated_at

        members.append({
            "subscription_id": sub.id,
            "counterparty_name": _member_client_name(sub),  # CONTACTS 2.0 Ф4: Company-first
            "cohort_start_date": (
                sub.impl_start_date.isoformat()
                if sub.impl_start_date
                else (sub.created_at.date().isoformat() if sub.created_at else None)
            ),
            "current_stage_code": stage_code,
            "current_stage_name": stage_names.get(sub.lifecycle_stage_id) if sub.lifecycle_stage_id else None,
            "churn_date": churn_dt.date().isoformat() if churn_dt else None,
            "is_churned": churn_dt is not None,
            "fee_actual": float(sub.fee_actual) if sub.fee_actual is not None else None,
            "is_active": sub.is_active,
        })

    return {
        "cohort_month": cohort_month,
        "count": len(members),
        "members": sorted(members, key=lambda m: m["counterparty_name"]),
    }


@router.post("/cohorts/export")
async def export_cohorts_xlsx(
    current_user: CurrentUser,
    session: Annotated[AsyncSession, Depends(get_session)],
    periods: int = Query(12, ge=1, le=36),
):
    """Excel-экспорт матрицы когортного retention + LTV.

    Возвращает .xlsx с:
    - Лист «Retention»: матрица Cohort × Offset с процентами retention
    - Заголовок строки: когорта (YYYY-MM) + размер, столбцы: +0 мес .. +N мес + Avg LTV
    - freeze_panes на A2 (закреплён заголовок)
    """
    # Scope: fail-CLOSED через scope_query (зеркал /registry и /cohorts).
    from app.services.access_control import scope_query as _sq
    stmt = select(ClientSubscription)
    stmt = await _sq(session, stmt, ClientSubscription, "subscription", current_user)
    subs = (await session.execute(stmt)).scalars().all()

    # stage codes для churn fallback
    stage_ids = {s.lifecycle_stage_id for s in subs if s.lifecycle_stage_id}
    stage_codes: dict[int, str] = {}
    if stage_ids:
        stages_rows = (await session.execute(
            select(PipelineStage).where(PipelineStage.id.in_(stage_ids))
        )).scalars().all()
        stage_codes = {s.id: s.code for s in stages_rows}

    sub_ids = [s.id for s in subs]
    churn_dates: dict[int, datetime] = {}
    if sub_ids:
        history_rows = (await session.execute(
            select(SubscriptionStageHistory).where(
                SubscriptionStageHistory.subscription_id.in_(sub_ids),
                SubscriptionStageHistory.to_stage_code == "C0",
            ).order_by(SubscriptionStageHistory.changed_at)
        )).scalars().all()
        for h in history_rows:
            if h.subscription_id not in churn_dates:
                churn_dates[h.subscription_id] = h.changed_at

    for sub in subs:
        if sub.id in churn_dates:
            continue
        if stage_codes.get(sub.lifecycle_stage_id) == "C0" and sub.updated_at:
            churn_dates[sub.id] = sub.updated_at

    sub_dicts = _build_sub_dicts(subs, churn_dates, stage_codes)
    cohorts_data = build_cohort_data(sub_dicts, periods)

    max_offset = periods
    headers, rows = build_cohort_xlsx_rows(cohorts_data, max_offset)

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Retention"
    ws.freeze_panes = "A2"

    # Заголовок
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="172747")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    # Данные
    for row in rows:
        ws.append(row)

    # Ширина колонок
    ws.column_dimensions["A"].width = 12  # Когорта
    ws.column_dimensions["B"].width = 8   # Размер
    for i in range(3, 3 + max_offset + 1):
        letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[letter].width = 9
    # Avg LTV — последняя колонка
    last_col = ws.cell(row=1, column=len(headers)).column_letter
    ws.column_dimensions[last_col].width = 16

    buf = BytesIO()
    wb.save(buf)
    content = buf.getvalue()

    return Response(
        content=content,
        media_type=_XLSX_MEDIA,
        headers={"Content-Disposition": 'attachment; filename="cohorts.xlsx"'},
    )


# ============ DEALS 2.0 Ф3 — отчёты по сделкам ============

# Вспомогательная функция: загрузить имена компаний и ответственных одним батч-запросом.

async def _load_company_names(
    session: AsyncSession,
    deals: list[Deal],
) -> dict[int, str]:
    """Загрузить имена компаний одним запросом. Company-first, fallback на Counterparty."""
    co_ids = {d.company_id for d in deals if d.company_id}
    cp_ids = {d.counterparty_id for d in deals if not d.company_id and d.counterparty_id}

    co_map: dict[int, str] = {}
    if co_ids:
        cos = (await session.execute(
            select(Company).where(Company.id.in_(co_ids))
        )).scalars().all()
        co_map = {c.id: (c.name or c.legal_name or "") for c in cos}

    cp_map: dict[int, str] = {}
    if cp_ids:
        cps = (await session.execute(
            select(Counterparty).where(Counterparty.id.in_(cp_ids))
        )).scalars().all()
        cp_map = {c.id: (c.name or "") for c in cps}

    result: dict[int, str] = {}
    for d in deals:
        if d.company_id:
            result[d.id] = co_map.get(d.company_id, "")
        elif d.counterparty_id:
            result[d.id] = cp_map.get(d.counterparty_id, "")
        else:
            result[d.id] = ""
    return result


async def _load_user_names(
    session: AsyncSession,
    user_ids: set[int],
) -> dict[int, str]:
    if not user_ids:
        return {}
    users = (await session.execute(
        select(User).where(User.id.in_(user_ids))
    )).scalars().all()
    return {u.id: (u.full_name or "") for u in users}


async def _get_sales_stage_map(
    session: AsyncSession,
) -> dict[str, PipelineStage]:
    """Вернуть {code: stage} для sales-воронки (первой активной)."""
    pipe = (await session.execute(
        select(Pipeline).where(
            Pipeline.kind == "sales",
            Pipeline.is_active.is_(True),
        ).order_by(Pipeline.id).limit(1)
    )).scalar_one_or_none()
    if not pipe:
        return {}
    stages = (await session.execute(
        select(PipelineStage).where(PipelineStage.pipeline_id == pipe.id)
    )).scalars().all()
    return {s.code: s for s in stages if s.code}


async def _apply_deal_scope(
    session: AsyncSession,
    stmt,
    current_user: User,
):
    """Применить scope-фильтр видимости сделок (как в /deals board)."""
    from app.services.access_control import scope_query
    return await scope_query(session, stmt, Deal, "deal", current_user)


@router.get("/deals/without-tasks")
async def deals_without_tasks(
    current_user: CurrentUser,
    session: Annotated[AsyncSession, Depends(get_session)],
):
    """Wave 2a: счётчик «Сделки без задач» для виджета дашборда.

    «Без задач» = ОТКРЫТАЯ сделка (этап не won/lost) без открытой task-like
    Activity (kind ∈ {task,call,meeting}, completed_at IS NULL, is_closed=False).
    Определение зеркалит /deals?no_tasks=true и pure-предикат
    analytics_deals.is_open_task_activity.

    Scoped по роли: scope_query (manager — свои сделки, admin/director — все).
    Возвращает {count}. Клик по числу на дашборде → /deals?no_tasks=true.
    """
    open_stage_subq = (
        select(PipelineStage.id)
        .where(
            PipelineStage.is_won.is_(False),
            PipelineStage.is_lost.is_(False),
        )
    )
    open_task_exists = (
        select(Activity.id)
        .where(
            Activity.target_type == "deal",
            Activity.target_id == Deal.id,
            Activity.kind.in_(tuple(TASK_LIKE_ACTIVITY_KINDS)),
            Activity.completed_at.is_(None),
            Activity.is_closed.is_(False),
        )
    )
    stmt = (
        select(func.count(Deal.id))
        .where(
            Deal.stage_id.in_(open_stage_subq),
            ~open_task_exists.exists(),
        )
    )
    stmt = await _apply_deal_scope(session, stmt, current_user)
    count = (await session.execute(stmt)).scalar_one() or 0
    return {"count": int(count)}


@router.get("/deals/awaiting-payment")
async def deals_awaiting_payment(
    current_user: CurrentUser,
    session: Annotated[AsyncSession, Depends(get_session)],
):
    """Сделки в подстатусе await_payment с ожидаемыми платежами по договору.

    Возвращает:
    {
        items: [{
            deal_id, company_name, company_id,
            deal_amount, deal_currency,
            owner_name, owner_user_id,
            payments: [{payment_id, payment_date, amount, currency, notes}]
        }],
        total_deals: int,
        aggregation: aggregate_awaiting_payment result (by_week/by_month/grand_total/primary_currency)
    }

    «Ожидаемые платежи» = contract_payments с payment_date >= today для
    договора, привязанного к сделке (Deal.contract_id). Если договора нет —
    items показывается без платежей (warning в frontend).
    """
    today = date.today()

    # Найти этап await_payment в sales-воронке
    stage_map = await _get_sales_stage_map(session)
    await_stage = stage_map.get("await_payment")
    if not await_stage:
        return {"items": [], "total_deals": 0, "aggregation": aggregate_awaiting_payment([], today=today)}

    # Сделки в этапе await_payment
    stmt = select(Deal).where(Deal.stage_id == await_stage.id)
    stmt = await _apply_deal_scope(session, stmt, current_user)
    deals = (await session.execute(stmt)).scalars().all()

    if not deals:
        return {"items": [], "total_deals": 0, "aggregation": aggregate_awaiting_payment([], today=today)}

    # Загрузить имена компаний и ответственных
    company_names = await _load_company_names(session, list(deals))
    owner_ids = {d.owner_user_id for d in deals if d.owner_user_id}
    owner_names = await _load_user_names(session, owner_ids)

    # Загрузить ожидаемые платежи (payment_date >= today) по договорам сделок
    contract_ids = {d.contract_id for d in deals if d.contract_id}
    future_payments: dict[int, list[ContractPayment]] = {}  # contract_id → payments
    if contract_ids:
        pmts = (await session.execute(
            select(ContractPayment).where(
                ContractPayment.contract_id.in_(contract_ids),
                ContractPayment.payment_date >= today,
            ).order_by(ContractPayment.payment_date)
        )).scalars().all()
        for p in pmts:
            future_payments.setdefault(p.contract_id, []).append(p)

    items = []
    all_payment_dicts: list[dict] = []

    for d in deals:
        pmts_for_deal = future_payments.get(d.contract_id, []) if d.contract_id else []
        payment_rows = [
            {
                "payment_id": p.id,
                "payment_date": p.payment_date,
                "amount": float(p.amount),
                "currency": p.currency,
                "notes": p.notes,
            }
            for p in pmts_for_deal
        ]
        # Для агрегации формируем dict с company_name / owner_name
        for p in pmts_for_deal:
            all_payment_dicts.append({
                "payment_date": p.payment_date,
                "amount": p.amount,
                "currency": p.currency,
                "deal_id": d.id,
                "company_name": company_names.get(d.id, ""),
                "owner_name": owner_names.get(d.owner_user_id, "") if d.owner_user_id else "",
            })

        items.append({
            "deal_id": d.id,
            "deal_title": d.title,
            "company_name": company_names.get(d.id, ""),
            "company_id": d.company_id,
            "deal_amount": float(d.amount) if d.amount is not None else None,
            "deal_currency": d.currency,
            "owner_name": owner_names.get(d.owner_user_id, "") if d.owner_user_id else "",
            "owner_user_id": d.owner_user_id,
            "expected_close_date": d.expected_close_date.isoformat() if d.expected_close_date else None,
            "has_contract": d.contract_id is not None,
            "payments": payment_rows,
        })

    # Сортировка: сначала сделки с ближайшим expected_close_date
    items.sort(key=lambda x: (x["expected_close_date"] is None, x["expected_close_date"] or ""))

    aggregation = aggregate_awaiting_payment(all_payment_dicts, today=today)
    return {
        "items": items,
        "total_deals": len(items),
        "aggregation": aggregation,
    }


@router.get("/deals/paid")
async def deals_paid(
    current_user: CurrentUser,
    session: Annotated[AsyncSession, Depends(get_session)],
    limit: int = Query(50, ge=1, le=200),
):
    """Сделки в подстатусе paid (завершённые с первой оплатой).

    Возвращает:
    {
        items: [{
            deal_id, company_name, company_id,
            deal_amount, deal_currency,
            owner_name,
            contract_number,
            first_payment_date, first_payment_amount, first_payment_currency,
            paid_at (stage_changed_at),
        }],
        total_deals: int,
        summary: compute_paid_summary result
    }

    Первая оплата = самая ранняя ContractPayment по договору сделки.
    Если договора или платежей нет — поля first_payment_* = null.
    """
    # Найти этап paid
    stage_map = await _get_sales_stage_map(session)
    paid_stage = stage_map.get("paid")
    if not paid_stage:
        return {"items": [], "total_deals": 0, "summary": compute_paid_summary([])}

    stmt = (
        select(Deal)
        .where(Deal.stage_id == paid_stage.id)
        .order_by(Deal.stage_changed_at.desc().nulls_last())
        .limit(limit)
    )
    stmt = await _apply_deal_scope(session, stmt, current_user)
    deals = (await session.execute(stmt)).scalars().all()

    if not deals:
        return {"items": [], "total_deals": 0, "summary": compute_paid_summary([])}

    company_names = await _load_company_names(session, list(deals))
    owner_ids = {d.owner_user_id for d in deals if d.owner_user_id}
    owner_names = await _load_user_names(session, owner_ids)

    # Загрузить контракты для номеров и первых платежей
    contract_ids = {d.contract_id for d in deals if d.contract_id}
    contract_numbers: dict[int, str] = {}
    first_payments: dict[int, ContractPayment] = {}  # contract_id → earliest payment

    if contract_ids:
        contracts = (await session.execute(
            select(Contract).where(Contract.id.in_(contract_ids))
        )).scalars().all()
        contract_numbers = {c.id: (c.number or "") for c in contracts}

        # Все платежи по этим договорам, сортируем по дате (earliest first)
        all_pmts = (await session.execute(
            select(ContractPayment).where(
                ContractPayment.contract_id.in_(contract_ids),
            ).order_by(ContractPayment.payment_date)
        )).scalars().all()
        for p in all_pmts:
            if p.contract_id not in first_payments:
                first_payments[p.contract_id] = p

    items = []
    paid_dicts: list[dict] = []

    for d in deals:
        fp = first_payments.get(d.contract_id) if d.contract_id else None
        item = {
            "deal_id": d.id,
            "deal_title": d.title,
            "company_name": company_names.get(d.id, ""),
            "company_id": d.company_id,
            "deal_amount": float(d.amount) if d.amount is not None else None,
            "deal_currency": d.currency,
            "owner_name": owner_names.get(d.owner_user_id, "") if d.owner_user_id else "",
            "owner_user_id": d.owner_user_id,
            "contract_number": contract_numbers.get(d.contract_id, "") if d.contract_id else None,
            "first_payment_date": fp.payment_date.isoformat() if fp else None,
            "first_payment_amount": float(fp.amount) if fp else None,
            "first_payment_currency": fp.currency if fp else None,
            "paid_at": d.stage_changed_at.isoformat() if d.stage_changed_at else None,
        }
        items.append(item)
        paid_dicts.append({
            "deal_id": d.id,
            "company_name": item["company_name"],
            "contract_number": item["contract_number"],
            "first_payment_date": fp.payment_date if fp else None,
            "first_payment_amount": fp.amount if fp else None,
            "first_payment_currency": fp.currency if fp else None,
            "deal_amount": d.amount,
            "currency": d.currency,
        })

    summary = compute_paid_summary(paid_dicts)
    return {
        "items": items,
        "total_deals": len(items),
        "summary": summary,
    }


@router.get("/deals/hot")
async def deals_hot_analytics(
    current_user: CurrentUser,
    session: Annotated[AsyncSession, Depends(get_session)],
    limit: int = Query(50, ge=1, le=200),
):
    """Сделки в этапе hot — прогноз дожатия.

    Возвращает:
    {
        items: [{
            deal_id, company_name, company_id,
            deal_amount, deal_currency,
            owner_name,
            expected_close_date,
            days_to_close,          -- отрицательное = просрочено
            is_overdue,
        }],
        total_deals: int,
        forecast: compute_hot_forecast result
    }

    Сортировка: сначала просроченные, затем по ближайшей дате закрытия,
    затем без даты.
    """
    today = date.today()

    # Найти этап hot (верхнеуровневый, не подстатус)
    stage_map = await _get_sales_stage_map(session)
    hot_stage = stage_map.get("hot")
    if not hot_stage:
        return {"items": [], "total_deals": 0, "forecast": compute_hot_forecast([], today=today)}

    stmt = (
        select(Deal)
        .where(Deal.stage_id == hot_stage.id)
        .order_by(Deal.expected_close_date.asc().nulls_last())
        .limit(limit)
    )
    stmt = await _apply_deal_scope(session, stmt, current_user)
    deals = (await session.execute(stmt)).scalars().all()

    if not deals:
        return {"items": [], "total_deals": 0, "forecast": compute_hot_forecast([], today=today)}

    company_names = await _load_company_names(session, list(deals))
    owner_ids = {d.owner_user_id for d in deals if d.owner_user_id}
    owner_names = await _load_user_names(session, owner_ids)

    items = []
    hot_dicts: list[dict] = []

    for d in deals:
        days_to_close: int | None = None
        if d.expected_close_date is not None:
            days_to_close = (d.expected_close_date - today).days

        item = {
            "deal_id": d.id,
            "deal_title": d.title,
            "company_name": company_names.get(d.id, ""),
            "company_id": d.company_id,
            "deal_amount": float(d.amount) if d.amount is not None else None,
            "deal_currency": d.currency,
            "owner_name": owner_names.get(d.owner_user_id, "") if d.owner_user_id else "",
            "owner_user_id": d.owner_user_id,
            "expected_close_date": d.expected_close_date.isoformat() if d.expected_close_date else None,
            "days_to_close": days_to_close,
            "is_overdue": days_to_close is not None and days_to_close < 0,
        }
        items.append(item)
        hot_dicts.append({
            "deal_id": d.id,
            "company_name": item["company_name"],
            "amount": d.amount,
            "currency": d.currency,
            "expected_close_date": d.expected_close_date,
            "owner_name": item["owner_name"],
            "days_to_close": days_to_close,
        })

    # Пересортировать: просроченные первыми (by days_to_close ASC nulls last)
    items.sort(key=lambda x: (
        x["days_to_close"] is None,   # None последними
        x["days_to_close"] if x["days_to_close"] is not None else 0,
    ))

    forecast = compute_hot_forecast(hot_dicts, today=today)
    return {
        "items": items,
        "total_deals": len(items),
        "forecast": forecast,
    }
