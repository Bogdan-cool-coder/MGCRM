"""Чистые функции аналитики онбординга (Эпик 17).

Все функции принимают Python-структуры (list[dict]) — без обращения к БД.
Это позволяет покрывать их unit-тестами без fixture'ов.
"""
from __future__ import annotations

from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any


# ============ Overview aggregators ============


def compute_completion_rate(assignments_total: int, completions_total: int) -> float:
    """Процент завершений от общего числа назначений. Возвращает 0..100."""
    if assignments_total <= 0:
        return 0.0
    return round(completions_total / assignments_total * 100, 1)


def compute_avg_completion_hours(
    pairs: list[tuple[datetime | None, datetime | None]],
) -> float:
    """Среднее время прохождения курса (assigned_at, completed_at) в часах.

    Пары с None пропускаются.
    """
    deltas = [
        (b - a).total_seconds() / 3600
        for a, b in pairs
        if a is not None and b is not None and b >= a
    ]
    return round(sum(deltas) / len(deltas), 2) if deltas else 0.0


def fill_daily_gaps_int(
    rows: list[tuple[date, int]],
    start: date,
    end: date,
) -> list[int]:
    """Заполнить time-series по дням, вернуть только значения (int).

    Пропущенные даты → 0. Используется для sparkline.
    """
    by_date: dict[date, int] = {d: v for d, v in rows}
    out: list[int] = []
    cur = start
    while cur <= end:
        out.append(by_date.get(cur, 0))
        cur += timedelta(days=1)
    return out


def fill_daily_gaps_dict(
    rows: list[tuple[date, int]],
    start: date,
    end: date,
) -> list[dict[str, Any]]:
    """Заполнить time-series по дням, вернуть list[{date, count}].

    Пропущенные даты → count=0. Используется для activity sparkline (90d).
    """
    by_date: dict[date, int] = {d: v for d, v in rows}
    out: list[dict[str, Any]] = []
    cur = start
    while cur <= end:
        out.append({"date": cur.isoformat(), "count": by_date.get(cur, 0)})
        cur += timedelta(days=1)
    return out


# ============ Hard questions ============


def sort_hard_questions(
    rows: list[dict[str, Any]],
    limit: int = 5,
) -> list[dict[str, Any]]:
    """Отсортировать вопросы по success_rate_pct ASC (сложнейшие — первыми).

    rows — list[dict]: question_id, text, course_id, course_name, lesson_id,
    lesson_name, total_attempts, correct_attempts.
    Добавляет success_rate_pct: float.
    Фильтрует только строки с total_attempts >= 5 (статистически значимые).
    """
    enriched = []
    for r in rows:
        total = int(r.get("total_attempts") or 0)
        if total < 5:
            continue
        correct = int(r.get("correct_attempts") or 0)
        rate = round(correct / total * 100, 1) if total > 0 else 0.0
        enriched.append({**r, "success_rate_pct": rate})

    # Сортируем по success_rate ASC, вторичный ключ — total DESC (чаще встречается)
    enriched.sort(key=lambda x: (x["success_rate_pct"], -x.get("total_attempts", 0)))
    return enriched[:limit]


# ============ Funnel ============

# Ключи шагов воронки и их пороги (percent from CourseProgress)
FUNNEL_STEPS = [
    ("assigned",   "Назначено",    None,  None),   # все назначения
    ("started",    "Начато",       0,     None),    # percent > 0
    ("half_done",  "Прошёл 50%",  50,    None),    # percent >= 50
    ("near_done",  "Прошёл 90%",  90,    None),    # percent >= 90
    ("completed",  "Завершил",    100,   None),    # status == completed OR percent == 100
]


def compute_funnel_steps(
    assignments: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Вычислить шаги воронки онбординга для курса.

    assignments — list[dict]:
        user_id, percent (int 0..100), status ('not_started'|'in_progress'|'completed')

    Возвращает 5 шагов со step_key, step_label, count, pct_of_total.
    Логика:
      - assigned: все
      - started: percent > 0
      - half_done: percent >= 50
      - near_done: percent >= 90
      - completed: status == 'completed' OR percent == 100
    """
    total = len(assignments)
    if total == 0:
        return [
            {
                "step_key": key,
                "step_label": label,
                "count": 0,
                "pct_of_total": 0.0,
            }
            for key, label, _, _ in FUNNEL_STEPS
        ]

    counts: dict[str, int] = {}
    counts["assigned"] = total
    counts["started"] = sum(1 for a in assignments if int(a.get("percent") or 0) > 0)
    counts["half_done"] = sum(
        1 for a in assignments if int(a.get("percent") or 0) >= 50
    )
    counts["near_done"] = sum(
        1 for a in assignments if int(a.get("percent") or 0) >= 90
    )
    counts["completed"] = sum(
        1 for a in assignments
        if a.get("status") == "completed" or int(a.get("percent") or 0) == 100
    )

    out: list[dict[str, Any]] = []
    for key, label, _, _ in FUNNEL_STEPS:
        cnt = counts[key]
        pct = round(cnt / total * 100, 1) if total > 0 else 0.0
        out.append({
            "step_key": key,
            "step_label": label,
            "count": cnt,
            "pct_of_total": pct,
        })
    return out


def classify_funnel_step(percent: int, status: str) -> str:
    """Определить, на каком шаге воронки находится пользователь.

    Возвращает step_key самого дальнего шага, которого достиг юзер.
    """
    if status == "completed" or percent == 100:
        return "completed"
    if percent >= 90:
        return "near_done"
    if percent >= 50:
        return "half_done"
    if percent > 0:
        return "started"
    return "assigned"


# ============ Team progress aggregator ============


def aggregate_team_progress(
    rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Агрегировать список assignment-строк в per-user итоги.

    rows — list[dict]:
        user_id, full_name, email, department_name,
        status ('not_started'|'in_progress'|'completed'|'overdue'),
        percent (int), deadline_at (datetime|None), last_activity_at (datetime|None)

    Возвращает одну строку на user_id с агрегатами.
    """
    # Группировка по user_id
    by_user: dict[int, list[dict[str, Any]]] = {}
    for r in rows:
        uid = r["user_id"]
        by_user.setdefault(uid, []).append(r)

    out: list[dict[str, Any]] = []
    for uid, user_rows in by_user.items():
        first = user_rows[0]
        total = len(user_rows)
        completed = sum(1 for r in user_rows if r.get("status") == "completed")
        in_progress = sum(1 for r in user_rows if r.get("status") == "in_progress")
        overdue = sum(1 for r in user_rows if r.get("status") == "overdue")
        avg_pct = sum(int(r.get("percent") or 0) for r in user_rows) / total if total else 0.0

        # last_activity_at — максимум среди не-None
        activities = [r.get("last_activity_at") for r in user_rows if r.get("last_activity_at")]
        last_activity = max(activities) if activities else None

        out.append({
            "user_id": uid,
            "full_name": first.get("full_name", ""),
            "email": first.get("email", ""),
            "department_name": first.get("department_name"),
            "assignments_total": total,
            "completed_count": completed,
            "in_progress_count": in_progress,
            "overdue_count": overdue,
            "progress_pct": round(avg_pct, 1),
            "last_activity_at": last_activity,
        })
    return out


# ============ Excel builders ============


def build_overview_xlsx(
    totals: dict[str, Any],
    sparkline: list[int],
    course_completions: list[dict[str, Any]],
) -> bytes:
    """Excel-файл с overview-метриками (2 листа: KPI + детализация по курсам)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws_kpi = wb.active
    ws_kpi.title = "KPI"

    # Заголовок
    ws_kpi["A1"] = "Метрика"
    ws_kpi["B1"] = "Значение"
    ws_kpi["A1"].font = Font(bold=True)
    ws_kpi["B1"].font = Font(bold=True)
    ws_kpi["A1"].fill = PatternFill("solid", fgColor="172747")
    ws_kpi["B1"].fill = PatternFill("solid", fgColor="172747")
    for cell in ("A1", "B1"):
        ws_kpi[cell].font = Font(bold=True, color="FFFFFF")

    kpi_rows = [
        ("Активных курсов", totals.get("courses_active", 0)),
        ("Назначений всего", totals.get("assignments_total", 0)),
        ("Завершений всего", totals.get("completions_total", 0)),
        ("% завершения", f"{totals.get('completion_rate_pct', 0.0):.1f}%"),
        ("Среднее время прохождения (ч)", totals.get("avg_completion_hours", 0.0)),
        ("Активных учеников за 30д", totals.get("active_learners_30d", 0)),
        ("Просроченных обязательных", totals.get("overdue_mandatory", 0)),
        ("Новых назначений за 30д", totals.get("assignments_new_30d", 0)),
    ]
    for label, val in kpi_rows:
        ws_kpi.append([label, val])

    ws_kpi.column_dimensions["A"].width = 38
    ws_kpi.column_dimensions["B"].width = 18
    ws_kpi.freeze_panes = "A2"

    # Sparkline данные — отдельный диапазон под таблицей
    ws_kpi.append([])
    ws_kpi.append(["Назначений по дням (последние 30д):"])
    ws_kpi.append(sparkline if sparkline else [0] * 30)

    # Лист 2: прохождения по курсам
    ws_courses = wb.create_sheet("Прохождения по курсам")
    headers = ["Курс", "Завершений"]
    ws_courses.append(headers)
    ws_courses["A1"].font = Font(bold=True)
    ws_courses["B1"].font = Font(bold=True)
    for row in course_completions:
        ws_courses.append([row.get("title", ""), row.get("completed", 0)])
    ws_courses.column_dimensions["A"].width = 40
    ws_courses.column_dimensions["B"].width = 14
    ws_courses.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_team_progress_xlsx(rows: list[dict[str, Any]]) -> bytes:
    """Excel с прогрессом команды."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Прогресс команды"

    headers = [
        "ФИО", "Email", "Отдел",
        "Назначено", "Завершено", "В процессе", "Просрочено", "Прогресс %",
        "Последняя активность",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for r in rows:
        last_act = r.get("last_activity_at")
        last_act_str = last_act.strftime("%Y-%m-%d %H:%M") if last_act else ""
        ws.append([
            r.get("full_name", ""),
            r.get("email", ""),
            r.get("department_name") or "",
            r.get("assignments_total", 0),
            r.get("completed_count", 0),
            r.get("in_progress_count", 0),
            r.get("overdue_count", 0),
            r.get("progress_pct", 0.0),
            last_act_str,
        ])

    col_widths = [30, 28, 22, 12, 12, 14, 12, 12, 20]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # Числовой формат для процента
    for row in ws.iter_rows(min_row=2, min_col=8, max_col=8):
        for cell in row:
            cell.number_format = "0.0"

    ws.freeze_panes = "A2"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_hard_questions_xlsx(rows: list[dict[str, Any]]) -> bytes:
    """Excel с топ сложных вопросов."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Сложные вопросы"

    headers = [
        "Вопрос", "Курс", "Урок",
        "Попыток всего", "Правильных", "% успеха",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for r in rows:
        ws.append([
            r.get("text", r.get("question_text", "")),
            r.get("course_name", r.get("course_title", "")),
            r.get("lesson_name", r.get("lesson_title", "")),
            r.get("total_attempts", 0),
            r.get("correct_attempts", 0),
            r.get("success_rate_pct", 0.0),
        ])

    col_widths = [60, 30, 30, 14, 14, 12]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
        for cell in row:
            cell.number_format = "0.0"

    ws.freeze_panes = "A2"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
